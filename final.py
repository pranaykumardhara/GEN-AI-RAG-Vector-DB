import streamlit as st 
from streamlit_chat import message
from transformers import pipeline
import pandas as pd
import pandas as pd
import time
from pymilvus import connections, FieldSchema, CollectionSchema, DataType, Collection, utility
import csv
import json
import random
from openai import OpenAI 
import openai
import time
from dotenv import load_dotenv 
import os
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import openpyxl




# CLUSTER_ENDPOINT="https://in03-9d5726dc40dff99.api.gcp-us-west1.zillizcloud.com" # Set your cluster endpoint
# TOKEN="daa8b202deaa0e4db571f0ac315cd03c9c186f557c6aba83d6f195619299522015a64a543a140c95d4c8bb63b02545d8b9ee21fc" # Set your token
FILE = './data/ticket.xlsx'  # Download it from https://www.kaggle.com/datasets/jealousleopard/goodreadsbooks and save it in the folder that holds your script.
COLLECTION_NAME = 'title'  # Collection name
DIMENSION = 1536  # Embeddings size
COUNT = 100  # How many titles to embed and insert.
OPENAI_ENGINE = 'text-embedding-ada-002'  # Which engine to use
openai.api_key = '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'  # Use your own Open AI API Key here

load_dotenv()
os.environ['OPENAI_API_KEY'] = os.getenv('OPENAI_API_KEY')


@st.cache_data
def doc_preprocessing(file):
    with open(file, newline='',encoding='utf-8') as f:
        reader=csv.reader(f, delimiter=',')
        for row in reader:
            yield row[1]


# @st.cache_resource
def get_db():
    connections.connect(
    alias='default',
    uri="https://in03-d5c1cf295c9a349.api.gcp-us-west1.zillizcloud.com", 
    token="b3810d1dbe8e5a58c03beb0469b3c94db19c88dd05e56e4b90ad765d920cc318adf473ed2ae3be6615eb33a7283cdb95781f6483"
    )


def embed(text):
    client = OpenAI()
    return client.embeddings.create(input = [text], model=OPENAI_ENGINE).data[0].embedding
    # return ''



def insert_db():
    get_db()
    try:
        from pymilvus import utility
        utility.drop_collection(COLLECTION_NAME)
    except:
        pass
    fields = [
        FieldSchema(name='id', dtype=DataType.INT64, descrition='Ids', is_primary=True, auto_id=True),
        FieldSchema(name='title', dtype=DataType.VARCHAR, description='Title texts', max_length=2500),
        FieldSchema(name='embedding', dtype=DataType.FLOAT_VECTOR, description='Embedding vectors', dim=DIMENSION)
    ]
    # 2. Build the schema
    schema = CollectionSchema(
        fields,
        description="Schema of Medium articles",
    )
    collection = Collection(
        name=COLLECTION_NAME, 
        description="Medium articles published between Jan and August in 2020 in prominent publications",
        schema=schema
    )
    index_params = {
        'index_type': 'IVF_FLAT',
        'metric_type': 'L2',
        'params': {'nlist': 1024}
    }
    # To name the index, do as follows:
    collection.create_index(
        field_name="embedding", 
        index_params=index_params,
    )
    workbook = openpyxl.load_workbook(FILE)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        text= str(row[0].value) + ' is ' + str(row[1].value) + " it's created on " + str(row[2].value) + " and updated on " + str(row[3].value) + ". it's Description is '" + str(row[7].value) + "'. the issue type is " + str(row[10].value) + " and creator is " + str(row[11].value) + ", it's last viewed on " + str(row[12].value) + ". Incident: Time to Assignment is " + str(row[13].value) + " and Division is " + str(row[14].value) + ", Urgency is " + str(row[15].value) + " , Operational categorization is " + str(row[16].value) + ", Last Comment is " + str(row[17].value) + ". it's Assignee Manager is " + str(row[18].value) + ", it's Request Sub-Type is " + str(row[19].value) + ", Summary is " + str(row[20].value) + '\n'
        ins=[ [text], [embed(text)]]  # Insert the title id, the title text, and the title embedding vector
        collection.insert(ins)
        print("instd............")
        time.sleep(21)

    # i=0
    # # Insert each title and its embedding
    # for idx, text in enumerate(random.sample(sorted(doc_preprocessing(FILE)), k=COUNT)):  # Load COUNT amount of random values from dataset
        # if text.startswith("OMEGAUP-28291") :
        #     pass
        # else:
        #     ins=[ [text], [embed(text)]]  # Insert the title id, the title text, and the title embedding vector
        #     collection.insert(ins)
        #     i=i+1
        #     time.sleep(21)
        # if i == 15 :
        #     break
    collection.load()

# insert_db()

@st.cache_resource
def getpipeline():
 return pipeline("question-answering", model='bert-large-uncased-whole-word-masking-finetuned-squad')

@st.cache_resource
def search_db(question):
    get_db()
    collection = Collection(
        name=COLLECTION_NAME
    )
    res = collection.query(
        expr="title not in ['0']",
        offset = 0,
        limit = 50, 
        output_fields = ["title"]
    )
    titleobj=[question]
    for obj in res:
        titleobj.append(obj['title'])

    print('titleobj..............',titleobj)



    vectorizer = TfidfVectorizer(stop_words='english')
    tfidf_matrix = vectorizer.fit_transform(titleobj)

    # TF-IDF Vectorization
    feature_names = vectorizer.get_feature_names_out()
    df = pd.DataFrame(tfidf_matrix.toarray(), columns=feature_names)
    # print(df)

    # Calculate cosine similarity between job description and CVs
    cosine_similarities = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:]).flatten()
    similarity_scores = list(cosine_similarities)
    # Output: [0.704, 0.233]
    print(max(similarity_scores))
    # Presenting matched CVs based on similarity scores
    titleobj.remove(question)
    best_matched_cv = titleobj[similarity_scores.index(max(similarity_scores))]

    print("best_matched_cv............",best_matched_cv)
    
    question_answerer = getpipeline()

    return question_answerer(question=question,context=best_matched_cv)['answer']
   



def display_conversation(history):
    for i in range(len(history["generated"])):
        message(history["past"][i], is_user=True, key=str(i) + "_user")
        message(history["generated"][i],key=str(i))

def main():
    # Initialize Streamlit app with a title
    #st.title("LLM Powered Chatbot")
     # Initialize Streamlit app with a title
    st.title("NOV HelpDeskAmigo")
    st.image("iconnov.png")
    

    # Get user input from text input
    user_input = st.text_input("", key="input")

    # Initialize session state for generated responses and past messages
    if "generated" not in st.session_state:
        st.session_state["generated"] = ["I am ready to help you"]
    if "past" not in st.session_state:
        st.session_state["past"] = ["Hey there!"]
        
    # Search the database for a response based on user input and update session state
    if user_input:
        st.session_state.past.append(user_input)
        response = search_db(user_input)
        st.session_state.generated.append(response)
        col1, col2, col8, col9 = st.columns([6, 6,1,1])
        with col8:
            with st.container():
                st.button(":thumbsup:",use_container_width=True)
        with col9:
            with st.container():
                st.button(":thumbsdown:",use_container_width=True)
        col8.write(' ')
        col9.write(' ')
    # Display conversation history using Streamlit messages
    if st.session_state["generated"]:
        display_conversation(st.session_state)
        st.session_state.clear()

if __name__ == "__main__":
    main()
    pass








